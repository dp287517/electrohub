#!/usr/bin/env python3
"""
Script de conversion de matrice PDF Siemens FC2060 vers Excel
Extrait les zones, équipements et niveaux d'alarme (AL1/AL2) en préservant la structure spatiale

Usage:
    python pdf_matrix_to_excel.py input.pdf output.xlsx
    python pdf_matrix_to_excel.py input.pdf  # Crée input_matrix.xlsx
"""

import sys
import re
from pathlib import Path
from typing import Dict, List, Tuple, Optional
from dataclasses import dataclass, field

try:
    import pdfplumber
except ImportError:
    print("Installation de pdfplumber...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "pdfplumber"])
    import pdfplumber

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installation de openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter


@dataclass
class Equipment:
    """Représente un équipement dans la matrice"""
    address: str
    name: str
    description: str = ""
    alarm_level: str = ""  # AL1 ou AL2
    zone_id: Optional[str] = None
    page: int = 0
    row_position: float = 0


@dataclass
class Zone:
    """Représente une zone de détection incendie"""
    id: str
    name: str
    building: str = ""
    floor: str = ""
    location: str = ""
    equipment: List[Equipment] = field(default_factory=list)
    al1_count: int = 0
    al2_count: int = 0


class FC2060MatrixParser:
    """Parser pour les matrices cause-effet Siemens FC2060"""

    # Patterns de détection
    ZONE_PATTERNS = [
        r'Zone\s*[:\-]?\s*(\d+[\.\-]?\d*[\.\-]?\d*)',
        r'Z\.?\s*(\d+[\.\-]?\d*)',
        r'Boucle\s*(\d+)',
        r'Loop\s*(\d+)',
    ]

    AL_PATTERNS = [
        (r'\bAL\s*1\b|\bAL1\b|\bAlarme\s*1\b|\bAlarm\s*1\b|\bLocale\b', 'AL1'),
        (r'\bAL\s*2\b|\bAL2\b|\bAlarme\s*2\b|\bAlarm\s*2\b|\bG[ée]n[ée]rale\b', 'AL2'),
    ]

    EQUIPMENT_PATTERNS = [
        r'(\d+[\.\-]\d+[\.\-]\d+)',  # 20.9.09 ou 20-9-09
        r'DET\s*[\-:]?\s*(\w+)',      # DET-001
        r'DAD\s*[\-:]?\s*(\w+)',      # DAD-001
        r'DI\s*[\-:]?\s*(\w+)',       # Détecteur ionique
        r'DO\s*[\-:]?\s*(\w+)',       # Détecteur optique
        r'BM\s*[\-:]?\s*(\w+)',       # Bouton manuel
        r'SIR\s*[\-:]?\s*(\w+)',      # Sirène
    ]

    def __init__(self, pdf_path: str):
        self.pdf_path = Path(pdf_path)
        self.zones: Dict[str, Zone] = {}
        self.equipment_list: List[Equipment] = []
        self.current_zone: Optional[Zone] = None
        self.current_al: str = ""
        self.matrix_data: List[List[str]] = []

    def extract_building_from_name(self, name: str) -> str:
        """Extrait le bâtiment depuis le nom ou l'adresse"""
        # Pattern numérique type 20.9.09 -> B20
        match = re.search(r'^(\d+)[\.\-]', name)
        if match:
            return f"B{match.group(1)}"

        # Patterns textuels
        patterns = [
            (r'B[âa]t(?:iment)?\s*[:\-]?\s*(\w+)', r'B\1'),
            (r'Building\s*[:\-]?\s*(\w+)', r'B\1'),
            (r'\b(B\d+)\b', r'\1'),
        ]
        for pattern, replacement in patterns:
            match = re.search(pattern, name, re.IGNORECASE)
            if match:
                return re.sub(pattern, replacement, match.group(0), flags=re.IGNORECASE)

        return ""

    def detect_alarm_level(self, text: str) -> str:
        """Détecte le niveau d'alarme dans le texte"""
        text_upper = text.upper()
        for pattern, level in self.AL_PATTERNS:
            if re.search(pattern, text_upper, re.IGNORECASE):
                return level
        return ""

    def parse_zone_id(self, text: str) -> Optional[str]:
        """Extrait l'ID de zone du texte"""
        for pattern in self.ZONE_PATTERNS:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                return match.group(1)
        return None

    def extract_tables_from_page(self, page) -> List[List[List[str]]]:
        """Extrait les tables d'une page PDF"""
        tables = page.extract_tables()
        return tables if tables else []

    def process_table_row(self, row: List[str], page_num: int, row_idx: int):
        """Traite une ligne de tableau"""
        if not row or all(not cell for cell in row):
            return

        row_text = ' '.join(str(cell) for cell in row if cell)

        # Détecter changement de zone
        zone_id = self.parse_zone_id(row_text)
        if zone_id:
            if zone_id not in self.zones:
                zone_name = row_text.strip()
                building = self.extract_building_from_name(zone_name)
                self.zones[zone_id] = Zone(
                    id=zone_id,
                    name=zone_name,
                    building=building
                )
            self.current_zone = self.zones[zone_id]

        # Détecter niveau d'alarme
        al_level = self.detect_alarm_level(row_text)
        if al_level:
            self.current_al = al_level

        # Chercher des équipements dans la ligne
        for pattern in self.EQUIPMENT_PATTERNS:
            matches = re.findall(pattern, row_text, re.IGNORECASE)
            for match in matches:
                equipment = Equipment(
                    address=match,
                    name=row_text[:100],
                    alarm_level=self.current_al or self.detect_alarm_level(row_text),
                    zone_id=self.current_zone.id if self.current_zone else None,
                    page=page_num,
                    row_position=row_idx
                )
                self.equipment_list.append(equipment)

                if self.current_zone:
                    self.current_zone.equipment.append(equipment)
                    if equipment.alarm_level == 'AL1':
                        self.current_zone.al1_count += 1
                    elif equipment.alarm_level == 'AL2':
                        self.current_zone.al2_count += 1

    def extract_text_blocks(self, page) -> List[Tuple[str, float, float]]:
        """Extrait les blocs de texte avec leur position"""
        blocks = []
        chars = page.chars

        # Grouper les caractères par ligne (même y)
        lines: Dict[float, List] = {}
        for char in chars:
            y = round(char['top'], 1)
            if y not in lines:
                lines[y] = []
            lines[y].append(char)

        # Reconstruire le texte de chaque ligne
        for y, line_chars in sorted(lines.items()):
            line_chars.sort(key=lambda c: c['x0'])
            text = ''.join(c['text'] for c in line_chars)
            x = line_chars[0]['x0'] if line_chars else 0
            blocks.append((text.strip(), x, y))

        return blocks

    def analyze_matrix_structure(self, page) -> Dict:
        """Analyse la structure de la matrice cause-effet"""
        structure = {
            'zones': [],
            'causes': [],
            'effects': [],
            'matrix_cells': []
        }

        # Extraire le texte avec positions
        blocks = self.extract_text_blocks(page)

        # Identifier les en-têtes (première partie = zones/causes, haut = effets)
        for text, x, y in blocks:
            zone_id = self.parse_zone_id(text)
            if zone_id:
                structure['zones'].append({
                    'id': zone_id,
                    'text': text,
                    'x': x,
                    'y': y
                })

            al_level = self.detect_alarm_level(text)
            if al_level:
                structure['effects'].append({
                    'level': al_level,
                    'text': text,
                    'x': x,
                    'y': y
                })

        return structure

    def parse(self) -> Dict[str, Zone]:
        """Parse le PDF et extrait toutes les données"""
        print(f"Analyse du fichier: {self.pdf_path}")

        with pdfplumber.open(self.pdf_path) as pdf:
            total_pages = len(pdf.pages)
            print(f"Nombre de pages: {total_pages}")

            for page_num, page in enumerate(pdf.pages, 1):
                print(f"  Traitement page {page_num}/{total_pages}...")

                # Méthode 1: Extraction de tables
                tables = self.extract_tables_from_page(page)
                for table in tables:
                    for row_idx, row in enumerate(table):
                        self.process_table_row(row, page_num, row_idx)
                        self.matrix_data.append(row)

                # Méthode 2: Analyse de structure si pas de tables
                if not tables:
                    structure = self.analyze_matrix_structure(page)
                    # Traiter le texte brut de la page
                    text = page.extract_text() or ""
                    for line_idx, line in enumerate(text.split('\n')):
                        self.process_table_row([line], page_num, line_idx)

        print(f"\nRésultats:")
        print(f"  Zones trouvées: {len(self.zones)}")
        print(f"  Équipements trouvés: {len(self.equipment_list)}")

        return self.zones

    def export_to_excel(self, output_path: str):
        """Exporte les données vers un fichier Excel"""
        wb = openpyxl.Workbook()

        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        al1_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        al2_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # === Feuille 1: Résumé par Zone ===
        ws_zones = wb.active
        ws_zones.title = "Zones"

        zone_headers = ["Zone ID", "Nom", "Bâtiment", "Nb Équipements", "AL1", "AL2"]
        for col, header in enumerate(zone_headers, 1):
            cell = ws_zones.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_idx, zone in enumerate(sorted(self.zones.values(), key=lambda z: z.id), 2):
            data = [
                zone.id,
                zone.name,
                zone.building or "Non défini",
                len(zone.equipment),
                zone.al1_count,
                zone.al2_count
            ]
            for col, value in enumerate(data, 1):
                cell = ws_zones.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col == 5 and value > 0:  # AL1
                    cell.fill = al1_fill
                elif col == 6 and value > 0:  # AL2
                    cell.fill = al2_fill

        # Ajuster largeurs
        for col in range(1, len(zone_headers) + 1):
            ws_zones.column_dimensions[get_column_letter(col)].width = 20
        ws_zones.column_dimensions['B'].width = 50

        # === Feuille 2: Détail Équipements ===
        ws_equip = wb.create_sheet("Équipements")

        equip_headers = ["Adresse", "Zone", "Niveau Alarme", "Description", "Page"]
        for col, header in enumerate(equip_headers, 1):
            cell = ws_equip.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_idx, equip in enumerate(self.equipment_list, 2):
            data = [
                equip.address,
                equip.zone_id or "",
                equip.alarm_level,
                equip.name[:100],
                equip.page
            ]
            for col, value in enumerate(data, 1):
                cell = ws_equip.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col == 3:
                    if value == 'AL1':
                        cell.fill = al1_fill
                    elif value == 'AL2':
                        cell.fill = al2_fill

        for col in range(1, len(equip_headers) + 1):
            ws_equip.column_dimensions[get_column_letter(col)].width = 15
        ws_equip.column_dimensions['D'].width = 60

        # === Feuille 3: Matrice Cause-Effet ===
        ws_matrix = wb.create_sheet("Matrice")

        # Créer une matrice zone vs alarme
        zones_list = sorted(self.zones.values(), key=lambda z: z.id)

        # En-têtes
        matrix_headers = ["Zone", "Bâtiment", "AL1 (Locale)", "AL2 (Générale)", "Total"]
        for col, header in enumerate(matrix_headers, 1):
            cell = ws_matrix.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

        for row_idx, zone in enumerate(zones_list, 2):
            data = [
                zone.id,
                zone.building or "Non défini",
                zone.al1_count,
                zone.al2_count,
                zone.al1_count + zone.al2_count
            ]
            for col, value in enumerate(data, 1):
                cell = ws_matrix.cell(row=row_idx, column=col, value=value)
                cell.border = border
                if col == 3:
                    cell.fill = al1_fill
                elif col == 4:
                    cell.fill = al2_fill

        # Totaux
        total_row = len(zones_list) + 2
        ws_matrix.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True)
        ws_matrix.cell(row=total_row, column=3, value=sum(z.al1_count for z in zones_list))
        ws_matrix.cell(row=total_row, column=4, value=sum(z.al2_count for z in zones_list))
        ws_matrix.cell(row=total_row, column=5, value=sum(len(z.equipment) for z in zones_list))

        for col in range(1, 6):
            ws_matrix.column_dimensions[get_column_letter(col)].width = 18

        # === Feuille 4: Données brutes ===
        if self.matrix_data:
            ws_raw = wb.create_sheet("Données Brutes")
            for row_idx, row in enumerate(self.matrix_data, 1):
                for col_idx, cell_value in enumerate(row, 1):
                    ws_raw.cell(row=row_idx, column=col_idx, value=str(cell_value) if cell_value else "")

        # Sauvegarder
        wb.save(output_path)
        print(f"\nFichier Excel créé: {output_path}")

        # Stats finales
        print("\n" + "="*50)
        print("RÉSUMÉ DE L'EXTRACTION")
        print("="*50)
        print(f"Zones détectées:     {len(self.zones)}")
        print(f"Équipements totaux:  {len(self.equipment_list)}")

        total_al1 = sum(z.al1_count for z in self.zones.values())
        total_al2 = sum(z.al2_count for z in self.zones.values())
        print(f"Équipements AL1:     {total_al1}")
        print(f"Équipements AL2:     {total_al2}")

        # Liste des bâtiments
        buildings = set(z.building for z in self.zones.values() if z.building)
        if buildings:
            print(f"\nBâtiments détectés: {', '.join(sorted(buildings))}")


def main():
    if len(sys.argv) < 2:
        print("Usage: python pdf_matrix_to_excel.py <fichier.pdf> [sortie.xlsx]")
        print("\nExemple:")
        print("  python pdf_matrix_to_excel.py matrice_fc2060.pdf")
        print("  python pdf_matrix_to_excel.py matrice.pdf resultat.xlsx")
        sys.exit(1)

    pdf_path = sys.argv[1]

    if not Path(pdf_path).exists():
        print(f"Erreur: Le fichier '{pdf_path}' n'existe pas.")
        sys.exit(1)

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        output_path = Path(pdf_path).stem + "_matrix.xlsx"

    parser = FC2060MatrixParser(pdf_path)
    parser.parse()
    parser.export_to_excel(output_path)


if __name__ == "__main__":
    main()
